# encoding=utf-8
import numpy as num
import traceback
import talib as ta
import tushare as ts
import time
from datetime import datetime, timedelta
import logging
# 配置日志输出格式和级别
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger().setLevel(logging.INFO)

#######################################################################################################################
# ############################################################################################### 配置程序应用所需要环境PATH
import sys
import os

project_name = 'QKL0731'
rootPath = str(os.path.abspath(os.path.dirname(__file__)).split(project_name)[0]) + project_name
sys.path.append(rootPath)

curPath1 = os.path.abspath(os.path.dirname(__file__))
rootPath1 = os.path.split(curPath1)[0]
sys.path.append(rootPath1)
import common_image
import common
import common_mysqlUtil
import openpyxl
import common_notion


#######################################################################################################################
# ########################################################################################################## 跨域5周线策略
def strategy(zhouqi, endstr, database_id, hui_ce_yue_fen=time.strftime("%Y%m", time.localtime())):
    logging.info("策略入参，周期： %s,截止日期：%s，Notion数据库ID：%s，回测月份：%s", zhouqi, endstr, database_id, hui_ce_yue_fen)
    # 局部变量初始化
    count = 0

    ts.set_token('a0a3a3ee133d6623bf9072236a5a8423c1c021d00aba3eb0c7bdfa5e')
    pro = ts.pro_api()
    all_code = pro.stock_basic(exchange='', list_status='L', fields='ts_code,symbol,name,area,industry,list_date')
    # all_code = ts.get_stock_basics()
    all_code = all_code[1:-1].ts_code
    all_code_index_x = num.array(all_code)

    time_str = endstr
    fo_10 = open("主力_双均线10_" + zhouqi + "_" + time_str + ".txt", "w")
    fo_60 = open("主力_双均线60_" + zhouqi + "_" + time_str + ".txt", "w")
    fo_144 = open("主力_双均线144_" + zhouqi + "_" + time_str + ".txt", "w")

    url = 'https://hook.us1.make.com/r7gj5cb1go2l7x23i44tnyivdj7sy7ei'

    # 打开Excel文件
    workbook = openpyxl.load_workbook('定向增发_' + hui_ce_yue_fen + '.xlsx')
    # 选择工作表
    sheet_names = workbook.sheetnames
    logging.info("获取Excel中的sheet列表，%s", sheet_names)

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        # 遍历Excel文件
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                column1_value, column2_value, column3_value, column4_value, column5_value, \
                column6_value, column7_value, column8_value, column9_value, column10_value, column11_value = row
                codeItem = str(column1_value).strip('\n')
                mark = "发行价格为" + str(column6_value) + ",发行时间为" + str(column9_value)[:10] + ",锁定期为" + column11_value

                count = count + 1
                codeName = ''
                data = common_mysqlUtil.select_all_code_one(codeItem)
                if len(data) > 0:
                    codeName = data[0][1]
                logging.info("第%d个正在执行中：code为：%s,名称为：%s，excel行信息：%s", count, column1_value, codeName, str(row))

                data_history = ts.get_hist_data(codeItem, ktype=zhouqi, end=endstr)
                data_history = data_history.iloc[::-1]

                closeArray = num.array(data_history['close'])
                doubleCloseArray = num.asarray(closeArray, dtype='double')

                # 均线
                ma10 = ta.SMA(doubleCloseArray, timeperiod=10)
                sma10 = ta.EMA(ma10, timeperiod=10)

                ma60 = ta.SMA(doubleCloseArray, timeperiod=60)
                sma60 = ta.EMA(ma60, timeperiod=60)

                ma144 = ta.SMA(doubleCloseArray, timeperiod=144)
                sma144 = ta.EMA(ma144, timeperiod=144)

                if ma10[-1] > sma10[-1] and ma10[-2] < sma10[-2]:
                    logging.debug("双均线10：code为：%s,名称为：%s", column1_value, codeName)
                    fo_10.write(codeItem + "\n")
                    # data = {
                    #     's_code': codeItem,
                    #     's_name': codeName,
                    #     's_type': '10Day'
                    # }
                    # requests.post(url, data=data)

                    image_path, gai_nian = common_image.plt_image_tongyichutu_2(codeItem,
                                                                                "D",
                                                                                "【定向增发】双均线10",
                                                                                "【定向增发】双均线10", time_str, mark, 'multi')

                    image_url = "http://" + "8.218.97.91:8080" + "/" + image_path[6:]
                    common_notion.create_content(database_id=database_id, title=column2_value,
                                                 ce_lve_lei_xing='10天双均线金叉', tu_pian=image_url,
                                                 mark=mark, gai_nian=gai_nian, code=codeItem, create_time=time_str)

                if ma60[-1] > sma60[-1] and ma60[-2] < sma60[-2]:
                    logging.debug("双均线60：code为：%s,名称为：%s", column1_value, codeName)
                    fo_60.write(codeItem + "\n")
                    # data = {
                    #     's_code': codeItem,
                    #     's_name': codeName,
                    #     's_type': '60Day'
                    # }
                    # requests.post(url, data=data)

                    image_path, gai_nian = common_image.plt_image_tongyichutu_2(codeItem,
                                                                                "D",
                                                                                "【定向增发】双均线60",
                                                                                "【定向增发】双均线60", time_str, mark, 'multi')
                    image_url = "http://" + "8.218.97.91:8080" + "/" + image_path[6:]
                    common_notion.create_content(database_id=database_id, title=column2_value,
                                                 ce_lve_lei_xing='60天双均线金叉', tu_pian=image_url,
                                                 mark=mark, gai_nian=gai_nian, code=codeItem, create_time=time_str)

                if ma144[-1] > sma144[-1] and ma144[-2] < sma144[-2]:
                    logging.debug("双均线144：code为：%s,名称为：%s", column1_value, codeName)
                    fo_144.write(codeItem + "\n")
                    # data = {
                    #     's_code': codeItem,
                    #     's_name': codeName,
                    #     's_type': '144Day'
                    # }
                    # requests.post(url, data=data)

                    image_path, gai_nian = common_image.plt_image_tongyichutu_2(codeItem,
                                                                                "D",
                                                                                "【定向增发】双均线144",
                                                                                "【定向增发】双均线144", time_str, mark, 'multi')

                    image_url = "http://" + "8.218.97.91:8080" + "/" + image_path[6:]
                    common_notion.create_content(database_id=database_id, title=column2_value,
                                                 ce_lve_lei_xing='144天双均线金叉', tu_pian=image_url,
                                                 mark=mark, gai_nian=gai_nian, code=codeItem, create_time=time_str)

            except (IOError, TypeError, NameError, IndexError, Exception) as e:
                print(e)
                logging.error("AGU_主力_双均线执行异常,code为：%s,名称为：%s,excel行信息：%s,异常信息：%s",
                              column1_value, codeName, str(row), e)
                traceback.print_exc()
                common.dingding_markdown_msg_03("AGU_主力_双均线执行异常", "AGU_主力_双均线执行异常")

        # 关闭工作簿
        workbook.close()
    return ""


#######################################################################################################################
# ############################################################################################################# 主执行程序
if __name__ == "__main__":
    # 获取命令行参数数量（不包括脚本文件名）
    num_args = len(sys.argv) - 1
    if num_args < 1:
        logging.info("开始策略默认逻辑(当前时间)")
        # Notion数据库ID：任务跟踪（Auto）
        database_id = "163fe8f3baa744c2922f78657a7e7066"
        logging.info("清空Notion数据库，ID为：%s", database_id)
        common_notion.clear_database(database_id)

        time_str1 = time.strftime("%Y-%m-%d", time.localtime())
        strategy(zhouqi='D', endstr=time_str1, database_id=database_id)
        common.dingding_markdown_msg_03("AGU_主力_双均线执行完成", "AGU_主力_双均线执行完成")
    else:
        logging.info("开始策略回测逻辑")
        # Notion数据库ID：任务跟踪（Auto）（回测）
        database_id = "8ccf5196004949e386fa321384e6e76c"
        logging.info("清空Notion数据库，ID为：%s", database_id)
        common_notion.clear_database(database_id)

        # 按月进行回测,回测1月份数据
        # 设置起始日期为2023年1月1日
        start_date = datetime(2023, 1, 1)

        # 计算结束日期为2023年2月1日（不包含2月1日）
        end_date = datetime(2023, 2, 1)

        # 循环遍历日期
        current_date = start_date
        while current_date < end_date:
            time_str1 = current_date.strftime('%Y-%m-%d')
            time_str2 = current_date.strftime('%Y%m')
            strategy(zhouqi='D', endstr=time_str1, database_id=database_id, hui_ce_yue_fen=time_str2)
            current_date += timedelta(days=1)
