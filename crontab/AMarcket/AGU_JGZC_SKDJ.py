import numpy as num
import talib as ta
import tushare as ts
import time
import openpyxl

#######################################################################################################################
################################################################################################配置程序应用所需要环境PATH
import sys
import os

project_name = 'QKL0731'
rootPath = str(os.path.abspath(os.path.dirname(__file__)).split(project_name)[0]) + project_name
sys.path.append(rootPath)
import common
import common_image
import common_zhibiao

count = 0
count_b = 0
time_str = time.strftime("%Y%m%d", time.localtime())


#######################################################################################################################
####################################################################################################################EXE
def exe(one_column_data, zhouqi, endstr, fo):
    print('{:0>6d}'.format(one_column_data[1].value))
    print(one_column_data[2].value)
    codeItem = '{:0>6d}'.format(one_column_data[1].value)
    time.sleep(0.01)
    print("==========================================================" + codeItem)

    try:
        data_history = ts.get_hist_data(codeItem, ktype=zhouqi, end=endstr)
        data_history = data_history.iloc[::-1]
        closeArray = num.array(data_history['close'])
        doubleCloseArray = num.asarray(closeArray, dtype='double')

        highArray = num.array(data_history['high'])
        doubleHighArray = num.asarray(highArray, dtype='double')

        openArray = num.array(data_history['open'])
        doubleOpenArray = num.asarray(openArray, dtype='double')

        # print(data_history)
        k0, d0 = common_zhibiao.SKDJ_zhibiao(data_history, doubleCloseArray)

        if k0[-1]>k0[-2] and k0[-3]>k0[-2]:
            print(codeItem + "========================================")
            print(k0[-1])
            print(d0[-1])
            fo.write(codeItem + "==" + one_column_data[2].value + "==" + one_column_data[4].value + "\n")
            common.dingding_markdown_msg_2('触发【高瓴资本、社保基金机构新进】【'+ zhouqi +'】级别SKDJ翻转'
                                           + codeItem + "==" + one_column_data[2].value + "=="
                                           + one_column_data[4].value,
                                           '触发【高瓴资本、社保基金机构新进】【'+ zhouqi +'】级别SKDJ翻转'
                                           + codeItem + "==" + one_column_data[2].value + "=="
                                           + one_column_data[4].value)
            time.sleep(5)
        # if k0[-1] < 55 and k0[-2] < d0[-2] and k0[-1] > d0[-1]:
        #     print(codeItem + "========================================")
        #     print(k0[-1])
        #     print(d0[-1])
        #     MA_20 = ta.SMA(doubleCloseArray, timeperiod=20)
        #     if MA_20[-1] > MA_20[-2]:
        #         fo.write(codeItem + "\n")
    except (IOError, TypeError, NameError, IndexError, Exception) as e:
        print(e)

#######################################################################################################################
###########################################################################################################跨域5月线策略
def strategy(zhouqi, endstr):
    # 局部变量初始化
    fo = open("SKDJ_JGZC_" + zhouqi + "_" + endstr + ".txt", "w")
    fo1 = open("SKDJ_JGZC_" + zhouqi + "_" + endstr + ".txt", "w")
    fo2 = open("SKDJ_JGZC_" + zhouqi + "_" + endstr + ".txt", "w")
    fo3 = open("SKDJ_JGZC_" + zhouqi + "_" + endstr + ".txt", "w")
    fo4 = open("SKDJ_社保等新进" + zhouqi + "_" + endstr + ".txt", "w", encoding='UTF-8')

    wb = openpyxl.load_workbook('JGZC.xlsx')
    # 获取所有工作表名
    names = wb.sheetnames
    # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
    sheet = wb[names[0]]
    sheet1 = wb[names[1]]
    sheet2 = wb[names[2]]
    sheet3 = wb[names[3]]
    sheet4 = wb[names[4]]

    # for one_column_data in sheet.iter_rows():
    #     exe(one_column_data,zhouqi, endstr, fo)
    # for one_column_data in sheet1.iter_rows():
    #     exe(one_column_data,zhouqi, endstr, fo)
    # for one_column_data in sheet2.iter_rows():
    #     exe(one_column_data,zhouqi, endstr, fo)
    # for one_column_data in sheet3.iter_rows():
    #     exe(one_column_data,zhouqi, endstr, fo)
    for one_column_data in sheet4.iter_rows():
        exe(one_column_data, zhouqi, endstr, fo4)
    fo4.close()

    return count_b


#######################################################################################################################
##############################################################################################################主执行程序
time_str1 = time.strftime("%Y-%m-%d", time.localtime())
count_result_b = strategy('D', time_str1)
count_result_b = strategy('W', time_str1)
# bp = ByPy()
# timeStr1 = time.strftime("%Y%m%d", time.localtime())
# bp.mkdir(remotepath=timeStr1)
# bp.upload(localpath=rootPath + os.sep + "images" + os.sep + timeStr1, remotepath=timeStr1)
common.dingding_markdown_msg_2('触发【高瓴资本、社保基金机构新进】SKDJ金叉', '触发【高瓴资本、社保基金机构新进】SKDJ金叉')
