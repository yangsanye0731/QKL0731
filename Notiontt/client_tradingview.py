import time
from pywinauto import application
from pywinauto.mouse import click, double_click, right_click, move
from pywinauto.keyboard import send_keys
from pywinauto import findwindows
import random
from PIL import ImageGrab
from pathlib import Path
import ocr_util as ocr
import openpyxl

#######################################################################################################################
# ############################################################################################### 配置程序应用所需要环境PATH
import sys
import os

project_name = 'QKL0731'
rootPath = str(os.path.abspath(os.path.dirname(__file__)).split(project_name)[0]) + project_name
sys.path.append(rootPath)

curPath1 = os.path.abspath(os.path.dirname(__file__))
rootPath1 = os.path.split(curPath1)[0]
sys.path.append(rootPath1)
import common


def image(code):
    # 在全屏情况下点击搜索图标
    time.sleep(random.randint(1, 2))
    click(coords=(68, 58))  # 在 (100, 200) 坐标处点击
    time.sleep(random.randint(1, 2))
    # 在搜索栏中输入代号
    click(coords=(609, 295))
    # 在文本框中输入全选与删除
    send_keys("^a")
    time.sleep(round(random.uniform(0.2, 0.5), 1))
    send_keys("{DELETE}")
    time.sleep(round(random.uniform(0.2, 0.5), 1))
    # 输入股票代码，这里是【变量】
    send_keys(code, pause=round(random.uniform(0.2, 1), 1))
    time.sleep(round(random.uniform(0.2, 0.5), 1))
    send_keys("{ENTER}")
    time.sleep(2)
    # 截取整个屏幕
    # 指定要截取的区域的坐标和大小
    x1, y1, x2, y2 = 1600, 987, 1683, 1015
    # 使用 ImageGrab.grab() 截取屏幕上的区域
    screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))
    # 创建文件夹
    time_str = time.strftime("%Y-%m-%d", time.localtime())
    folder_path = "images\\" + time_str

    # 使用 Path() 创建文件夹
    path = Path(folder_path)
    path.mkdir(parents=True, exist_ok=True)
    absolute_path = path.resolve()

    image_file_path = os.path.join(absolute_path, code + ".png")
    # 保存截图为文件
    screenshot.save(image_file_path)
    text = ocr.ocr(image_file_path)
    print(text)
    # 关闭截图
    screenshot.close()
    time.sleep(1)
    return text


def send_dingding_msg(code, price, count):
    time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    common.dingding_markdown_msg_03(
        time_str + '触发自动卖出' + code + '当前价格：' + price + ' 数量：' + count,
        time_str + '触发自动卖出' + code + '当前价格：' + price + ' 数量：' + count)


def maximize(title_str):
    windows = findwindows.find_windows(title_re=title_str + "")
    if windows:
        app = application.Application().connect(handle=windows[0])
        # window_title = app.top_window().window_text()
        app.top_window().wrapper_object().maximize()


def minimize(title_str):
    windows = findwindows.find_windows(title_re=title_str + "")
    if windows:
        app = application.Application().connect(handle=windows[0])
        # window_title = app.top_window().window_text()
        app.top_window().wrapper_object().minimize()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        title_str = "399006"
        # 最大化窗口
        maximize(title_str)
        excel_file_path = 'C:\\Users\\yangj\\Desktop\\buy\\my.xlsx'
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active

        time_str1 = time.strftime("%Y年%m月%d日%H时%M分", time.localtime())
        sheet.cell(row=1, column=sheet.max_column + 1, value=time_str1)
        row_count = 2
        for row in sheet.iter_rows(min_row=2, values_only=True, max_col=3):
            column1_value, column2_value, column3_value = row
            code = str(column2_value)
            print(code)
            text = image(code)
            if text == 'y':
                text = '中立'
            sheet.cell(row=row_count, column=sheet.max_column, value=text)
            row_count = row_count + 1

        # 保存 Excel 文件
        workbook.save(excel_file_path)
        # 关闭工作簿
        workbook.close()
        # 最小化窗口
        minimize(title_str)
        time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        common.dingding_markdown_msg_03(
            time_str + '触发TradingView技术指标统计完成',
            time_str + '触发TradingView技术指标统计完成')
    else:
        print("=====")
